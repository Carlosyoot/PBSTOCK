import pandas as pd
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import os

from functions.events.InterfaceError.popup import Popup
documentos_path = os.path.expanduser("~/Documents")

pbstock_path = os.path.join(documentos_path, "PBSTOCK")

if not os.path.exists(pbstock_path):
    os.makedirs(pbstock_path)



def exportar_para_excel(ui, nome_arquivo="tabela_monitoramento.xlsx"):
   
    try:
        caminho_completo = os.path.join(pbstock_path, nome_arquivo)

        dados = []
        colunas = []

        for col in range(ui.tabela_monitoramento.columnCount()):
            colunas.append(ui.tabela_monitoramento.horizontalHeaderItem(col).text())

        for row in range(ui.tabela_monitoramento.rowCount()):
            linha = []
            for col in range(ui.tabela_monitoramento.columnCount()):
                item = ui.tabela_monitoramento.item(row, col)
                if item is not None:
                    linha.append(item.text())
                else:
                    linha.append("")  
            dados.append(linha)

        df = pd.DataFrame(dados, columns=colunas)

        with pd.ExcelWriter(caminho_completo, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Monitoramento', index=False)
            workbook = writer.book
            worksheet = writer.sheets['Monitoramento']

            col_widths = [17] * (len(colunas))  
            col_widths[-1] = 17.07  

            for i, width in enumerate(col_widths, start=1):
                col_letter = get_column_letter(i)
                worksheet.column_dimensions[col_letter].width = width

            cinza_claro = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

            borda = Border(left=Side(style='thin'),
                           right=Side(style='thin'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))

            for col_num, cell in enumerate(worksheet[1], start=1):
                cell.fill = cinza_claro
                cell.border = borda

            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = borda

        Popup(f"Tabela exportada para \n{caminho_completo}")

    except Exception as e:
        print(f"Erro ao exportar para Excel: {e}")
        Popup(f"Erro ao exportar para PDF: {e}")
        
def exportar_para_pdf(tabela, nome_arquivo="tabela_monitoramento.pdf"):
    try:
        caminho_completo = os.path.join(pbstock_path, nome_arquivo)

        dados = [[tabela.horizontalHeaderItem(col).text() for col in range(tabela.columnCount())]]
        for row in range(tabela.rowCount()):
            linha = [tabela.item(row, col).text() if tabela.item(row, col) else "" for col in range(tabela.columnCount())]
            dados.append(linha)

        pdf_document = SimpleDocTemplate(caminho_completo, pagesize=A4)  
        estilo_tabela = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ])
        tabela_pdf = Table(dados)
        tabela_pdf.setStyle(estilo_tabela)
        pdf_document.build([tabela_pdf])

        Popup(f"Tabela exportada para \n{caminho_completo}")

    except Exception as e:
        print(f"Erro ao exportar para PDF: {e}")
        Popup(f"Erro ao exportar para PDF: {e}")